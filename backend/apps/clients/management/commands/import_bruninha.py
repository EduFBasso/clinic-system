"""Management command: importa pacientes do ERP odontológico da Dra. Bruna.

Uso:
    python manage.py import_bruninha --file /caminho/para/extracao_dados_bruninha.xlsx \
                                     --professional contato.drabrunacarvalho@gmail.com

Flags:
    --dry-run   Simula sem gravar no banco.
    --update    Atualiza registros já existentes (busca por phone ou email).
"""
import re
from datetime import date, datetime

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from apps.clients.models import Client
from apps.register.models import Professional


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _split_name(full: str):
    """Divide 'Jose Ricardo Miguel' em first_name='Jose' e last_name='Ricardo Miguel'."""
    parts = (full or "").strip().split(None, 1)
    if len(parts) == 2:
        return parts[0], parts[1]
    return (parts[0] if parts else ""), ""


def _map_sex(tx_sexo: str):
    mapping = {
        "masculino": "masculino",
        "feminino": "feminino",
    }
    return mapping.get((tx_sexo or "").strip().lower(), None)


def _map_marital(tx_estado: str):
    mapping = {
        "solteiro": "solteiro",
        "solteira": "solteiro",
        "casado": "casado",
        "casada": "casado",
        "divorciado": "divorciado",
        "divorciada": "divorciado",
        "viúvo": "viuvo",
        "viúva": "viuvo",
        "viuvo": "viuvo",
        "viuva": "viuvo",
        "união estável": "uniao_estavel",
        "uniao estavel": "uniao_estavel",
    }
    return mapping.get((tx_estado or "").strip().lower(), None)


def _clean_cpf(tx_cpf: str):
    """Retorna apenas dígitos do CPF, ou None se inválido."""
    if not tx_cpf:
        return None
    digits = re.sub(r"\D", "", str(tx_cpf))
    return digits if len(digits) == 11 else None


def _parse_date(value) -> date | None:
    """Converte datetime do Excel ou string para date. Ignora placeholder 1800."""
    if not value:
        return None
    if isinstance(value, datetime):
        if value.year <= 1900:
            return None
        return value.date()
    if isinstance(value, date):
        return value if value.year > 1900 else None
    return None


def _clean_phone(tx_fone_label: str):
    """Normaliza telefone para formato E.164 brasileiro: +5519999999999."""
    if not tx_fone_label:
        return None
    digits = re.sub(r"\D", "", str(tx_fone_label))
    if not digits:
        return None
    if not digits.startswith("55"):
        digits = "55" + digits
    return "+" + digits


def _clean_postal(cep: str):
    if not cep:
        return None
    return re.sub(r"\D", "", str(cep))


# ---------------------------------------------------------------------------
# Main command
# ---------------------------------------------------------------------------

class Command(BaseCommand):
    help = "Importa pacientes do ERP odontológico da Dra. Bruna (arquivo .xlsx)."

    def add_arguments(self, parser):
        parser.add_argument(
            "--file", required=True,
            help="Caminho para o arquivo extracao_dados_bruninha.xlsx"
        )
        parser.add_argument(
            "--professional", required=True,
            help="E-mail do profissional que será vinculado aos clientes."
        )
        parser.add_argument(
            "--dry-run", action="store_true", default=False,
            help="Simula a importação sem gravar no banco."
        )
        parser.add_argument(
            "--update", action="store_true", default=False,
            help="Atualiza clientes já existentes (busca por telefone ou e-mail)."
        )

    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError:
            raise CommandError("openpyxl não instalado. Execute: pip install openpyxl")

        file_path = options["file"]
        prof_email = options["professional"]
        dry_run = options["dry_run"]
        do_update = options["update"]

        # Locate professional
        try:
            professional = Professional.objects.get(email__iexact=prof_email)
        except Professional.DoesNotExist:
            raise CommandError(f"Profissional '{prof_email}' não encontrado no banco.")

        self.stdout.write(f"Profissional: {professional.first_name} {professional.last_name} ({professional.email})")

        # Load workbook
        self.stdout.write(f"Abrindo: {file_path} ...")
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

        # ---------------------------------------------------------------
        # Build lookup tables from auxiliary sheets (keyed by ID_PACIENTE)
        # ---------------------------------------------------------------
        self.stdout.write("Carregando e-mails ...")
        emails = {}  # id_paciente -> email
        ws_email = wb["PACIENTE_EMAIL"]
        hdr_email = [c.value for c in next(ws_email.iter_rows(max_row=1))]
        for row in ws_email.iter_rows(min_row=2, values_only=True):
            r = dict(zip(hdr_email, row))
            if r.get("FL_ATIVO") and r.get("TX_EMAIL") and r.get("ID_PACIENTE"):
                pid = r["ID_PACIENTE"]
                if pid not in emails:
                    emails[pid] = str(r["TX_EMAIL"]).strip().lower()

        self.stdout.write("Carregando telefones ...")
        phones = {}  # id_paciente -> phone
        ws_fone = wb["PACIENTE_FONE"]
        hdr_fone = [c.value for c in next(ws_fone.iter_rows(max_row=1))]
        for row in ws_fone.iter_rows(min_row=2, values_only=True):
            r = dict(zip(hdr_fone, row))
            if r.get("FL_ATIVO") and r.get("TX_FONE_LABEL") and r.get("ID_PACIENTE"):
                pid = r["ID_PACIENTE"]
                if pid not in phones:
                    phones[pid] = _clean_phone(str(r["TX_FONE_LABEL"]))

        self.stdout.write("Carregando endereços ...")
        addresses = {}  # id_paciente -> dict
        ws_end = wb["PACIENTE_ENDERECO"]
        hdr_end = [c.value for c in next(ws_end.iter_rows(max_row=1))]
        for row in ws_end.iter_rows(min_row=2, values_only=True):
            r = dict(zip(hdr_end, row))
            if r.get("FL_ATIVO") and r.get("ID_PACIENTE"):
                pid = r["ID_PACIENTE"]
                if pid not in addresses:
                    addresses[pid] = {
                        "address": str(r.get("TX_ENDERECO") or "").strip() or None,
                        "address_number": str(r.get("TX_NUMERO") or "").strip() or None,
                        "address_complement": str(r.get("TX_COMPLEMENTO") or "").strip() or None,
                        "neighborhood": str(r.get("TX_BAIRRO") or "").strip() or None,
                        "city": str(r.get("TX_CIDADE") or "").strip() or None,
                        "state": str(r.get("CD_UF") or r.get("TX_UF") or "").strip()[:2] or None,
                        "postal_code": _clean_postal(r.get("TX_CEP")),
                    }

        # ---------------------------------------------------------------
        # Process PACIENTE_DADOS
        # ---------------------------------------------------------------
        self.stdout.write("Processando PACIENTE_DADOS ...")
        ws_pac = wb["PACIENTE_DADOS"]
        hdr_pac = [c.value for c in next(ws_pac.iter_rows(max_row=1))]

        created = updated = skipped = errors = 0

        with transaction.atomic():
            for row in ws_pac.iter_rows(min_row=2, values_only=True):
                r = dict(zip(hdr_pac, row))

                # Skip inactive patients
                if not r.get("FL_ATIVO"):
                    skipped += 1
                    continue

                pid = r.get("ID_PACIENTE")
                full_name = str(r.get("TX_NOME") or "").strip()
                if not full_name:
                    skipped += 1
                    continue

                first_name, last_name = _split_name(full_name)
                email = emails.get(pid)
                phone = phones.get(pid)
                addr = addresses.get(pid, {})

                # Build observations (merge TX_OBSERV fields)
                obs_parts = [
                    str(r.get(f"TX_OBSERV{s}") or "").strip()
                    for s in ["", "_CONTINUACAO1", "_CONTINUACAO2",
                              "_CONTINUACAO3", "_CONTINUACAO4", "_CONTINUACAO5"]
                ]
                clinical_history = "\n".join(p for p in obs_parts if p) or None

                # CPF
                cpf = _clean_cpf(r.get("TX_CPF"))

                data = {
                    "professional": professional,
                    "first_name": first_name,
                    "last_name": last_name,
                    "email": email or None,
                    "phone": phone,
                    "sex": _map_sex(r.get("TX_SEXO")),
                    "marital_status": _map_marital(r.get("TX_ESTADO_CIVIL")),
                    "date_of_birth": _parse_date(r.get("DT_NASCIMENTO")),
                    "document_type": "cpf" if cpf else None,
                    "document_number": cpf,
                    "profession": str(r.get("TX_OCUPACAO") or "").strip() or None,
                    "clinical_history": clinical_history,
                    **addr,
                }

                if dry_run:
                    created += 1
                    continue

                try:
                    # Try to find existing by phone or email
                    existing = None
                    if do_update:
                        if phone:
                            existing = Client.objects.filter(
                                professional=professional, phone=phone
                            ).first()
                        if not existing and email:
                            existing = Client.objects.filter(
                                professional=professional, email=email
                            ).first()

                    if existing:
                        for k, v in data.items():
                            if v is not None:
                                setattr(existing, k, v)
                        existing.save()
                        updated += 1
                    else:
                        # Avoid duplicate phone/email across all professionals (unique constraint)
                        if phone and Client.objects.filter(phone=phone).exists():
                            data["phone"] = None
                        if email and Client.objects.filter(email=email).exists():
                            data["email"] = None
                        Client.objects.create(**data)
                        created += 1

                except Exception as exc:
                    errors += 1
                    self.stderr.write(f"  ERRO pid={pid} nome='{full_name}': {exc}")

        mode = "[DRY-RUN] " if dry_run else ""
        self.stdout.write(self.style.SUCCESS(
            f"\n{mode}Concluído: {created} criados | {updated} atualizados | "
            f"{skipped} ignorados | {errors} erros"
        ))
