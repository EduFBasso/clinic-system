import openpyxl

wb = openpyxl.load_workbook(
    '/Users/eduardofigueiredobasso/Documents/clinic-system/clinic-system/extracao_dados_bruninha.xlsx',
    read_only=True
)

for name in wb.sheetnames:
    ws = wb[name]
    header = next(ws.iter_rows(max_row=1, values_only=True), ())
    cols = [str(h) for h in header if h is not None]
    print(f"{name}: {', '.join(cols)}")
